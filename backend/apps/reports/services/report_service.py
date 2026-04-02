"""
Richwell Portal — Report Service

Centralized logic for generating academic reports, Excel exports, and PDF documents.
Handles complex GPA calculations, curriculum audits, and dashboard statistics.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from django.db.models import Sum, Count
from apps.students.models import Student, StudentEnrollment
from apps.grades.models import Grade
from apps.scheduling.models import Schedule
from apps.academics.models import Subject, Program
from apps.facilities.models import Room
from apps.faculty.models import Professor
from apps.finance.models import Payment
from apps.terms.models import Term
from apps.auditing.models import AuditLog
from django.utils import timezone

class ReportService:
    """
    Service for all report-related business logic and document generation.
    """

    @staticmethod
    def generate_masterlist_excel(term_id, program_id=None, year_level=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Masterlist"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ["ID Number", "Last Name", "First Name", "Middle Name", "Program", "Year", "Gender", "Status"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, Alignment(horizontal="center"), border

        enrollments = StudentEnrollment.objects.filter(term_id=term_id)
        if program_id: enrollments = enrollments.filter(student__program_id=program_id)
        if year_level: enrollments = enrollments.filter(year_level=year_level)

        for row_num, enrollment in enumerate(enrollments, 2):
            s = enrollment.student
            vals = [s.idn, s.user.last_name, s.user.first_name, s.middle_name or "", s.program.code, f"Year {enrollment.year_level}", s.get_gender_display(), s.get_status_display()]
            for i, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=i, value=v)
                cell.border = border

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_cor_pdf(student_id, term_id):
        student, enrollment = Student.objects.get(id=student_id), StudentEnrollment.objects.get(student_id=student_id, term_id=term_id)
        grades = Grade.objects.filter(student_id=student_id, term_id=term_id, advising_status='APPROVED')
        if not grades.exists():
            raise ValueError("No approved subjects found for this student in the selected term. Ensure advising is complete and approved.")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch)
        elements, styles = [], getSampleStyleSheet()
        elements.append(Paragraph("<b>RICHWELL COLLEGES, INC.</b>", ParagraphStyle('Title', alignment=1, fontSize=14, spaceAfter=20)))
        
        # Student Info
        data = [[f"ID: {student.idn}", f"NAME: {student.user.get_full_name().upper()}"], [f"PROG: {student.program.code}", f"TERM: {enrollment.term.code}"]]
        elements.append(Table(data, colWidths=[2.5*inch, 4.5*inch]))
        elements.append(Spacer(1, 0.2*inch))

        # Subjects
        rows = [["CODE", "DESCRIPTION", "UNITS", "SCHEDULE", "ROOM"]]
        for g in grades:
            sch = Schedule.objects.filter(subject=g.subject, term_id=term_id).first()
            rows.append([g.subject.code, Paragraph(g.subject.description, styles['Normal']), str(g.subject.total_units), f"{''.join(sch.days)} {sch.start_time.strftime('%I:%M%p')}" if sch else "TBA", sch.room.name if sch and sch.room else "TBA"])
        
        t = Table(rows, colWidths=[1*inch, 2.5*inch, 0.6*inch, 2*inch, 1*inch])
        t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F172A')), ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke), ('GRID', (0,0), (-1,-1), 0.5, colors.grey)]))
        elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def get_academic_summary(student_id):
        student = Student.objects.select_related('user', 'program', 'curriculum').get(id=student_id)
        enrollment = student.enrollments.order_by('-enrollment_date').first()
        grades = Grade.objects.filter(student=student).select_related('subject', 'term')
        
        passed = grades.filter(grade_status='PASSED')
        units_earned = passed.aggregate(total=Sum('subject__total_units'))['total'] or 0
        numeric = grades.filter(final_grade__isnull=False)
        gpa = round(sum(g.final_grade * g.subject.total_units for g in numeric) / sum(g.subject.total_units for g in numeric), 2) if numeric.exists() else 0
        
        # Semester grouping
        semesters = []
        cur_subs = Subject.objects.filter(curriculum=student.curriculum).order_by('year_level', 'semester')
        grade_map = {g.subject_id: g for g in grades}
        
        for year in range(1, 6):
            for sem in ['1', '2', 'S']:
                subs = cur_subs.filter(year_level=year, semester=sem)
                if not subs.exists(): continue
                
                sem_grades = []
                for s in subs:
                    g = grade_map.get(s.id)
                    sem_grades.append({
                        "code": s.code, "subject": s.description, "units": s.total_units,
                        "grade": str(g.final_grade) if g and g.final_grade else (g.get_grade_status_display() if g else "--"),
                        "status": g.get_grade_status_display() if g else "Not Taken",
                        "status_code": g.grade_status if g else "NOT_TAKEN"
                    })
                sem_title = f"{'1st' if sem == '1' else '2nd' if sem == '2' else 'Summer'} Semester"
                semesters.append({
                    "title": sem_title, 
                    "year_level": f"Year {year}", 
                    "grades": sem_grades
                })

        return {
            "student": {
                "name": student.user.get_full_name(), 
                "idn": student.idn,
                "program": student.program.code, 
                "year_level": enrollment.year_level if enrollment else 1,
                "academic_standing": student.get_status_display(),
                "units_earned": units_earned
            },
            "stats": {"gpa": gpa, "passed": passed.count()},
            "semesters": semesters,
            "curriculum_progress": [] # Placeholder or real progress if needed
        }

    @staticmethod
    def get_dashboard_stats(user):
        role = user.role
        if role == 'ADMIN':
            return {"programs": Program.objects.count(), "professors": Professor.objects.count(), "rooms": Room.objects.count(), "audit_count": AuditLog.objects.count()}
        if role == 'REGISTRAR':
            return {"pending_docs": Student.objects.filter(status='APPLICANT').count(), "pending_advising": StudentEnrollment.objects.filter(advising_status='PENDING').count(), "total_students": Student.objects.count()}
        if role == 'CASHIER':
            return {"today": Payment.objects.filter(created_at__date=timezone.now().date()).aggregate(s=Sum('amount'))['s'] or 0, "pending_promissories": Payment.objects.filter(is_promissory=True).count()}
        
        if role == 'STUDENT':
            # Check if student is active for current term
            active_term = Term.objects.filter(is_active=True).first()
            student = user.student_profile
            enrollment = student.enrollments.filter(term=active_term).first()
            
            return {
                "active_term": {
                    "id": active_term.id,
                    "code": active_term.code,
                    "name": f"{active_term.get_semester_type_display()} {active_term.academic_year}",
                    "enrollment_open": active_term.is_active # Simplified, could use real dates here
                } if active_term else None,
                "current_enrollment": {
                    "id": enrollment.id,
                    "advising_status": enrollment.advising_status,
                    "is_regular": enrollment.is_regular
                } if enrollment else None,
                "student_info": {
                    "id": student.id,
                    "idn": student.idn,
                    "program": student.program.code,
                    "status": student.status
                }
            }
        return {}

    @staticmethod
    def graduation_check(student_id):
        student = Student.objects.get(id=student_id)
        req = Subject.objects.filter(curriculum=student.curriculum)
        passed_ids = set(Grade.objects.filter(student=student, grade_status='PASSED').values_list('subject_id', flat=True))
        missing = req.exclude(id__in=passed_ids)
        return {"is_eligible": not missing.exists(), "earned": sum(g.subject.total_units for g in Grade.objects.filter(student=student, grade_status='PASSED')), "required": sum(s.total_units for s in req), "missing": [{"code": m.code, "name": m.description} for m in missing]}
