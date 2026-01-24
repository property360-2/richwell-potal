"""
Export Service - Generate Excel and PDF exports from data.
"""

from io import BytesIO
from typing import List, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas


class ExportService:
    """
    Service for exporting data to Excel and PDF formats.
    """
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], columns: List[Dict[str, str]], title: str = "Export") -> BytesIO:
        """
        Export data to Excel format.
        
        Args:
            data: List of dictionaries containing row data
            columns: List of dicts with 'key' and 'label' for each column
            title: Sheet title
            
        Returns:
            BytesIO: Excel file in memory
            
        Example:
            columns = [
                {'key': 'id', 'label': 'ID'},
                {'key': 'name', 'label': 'Student Name'},
            ]
            data = [
                {'id': 1, 'name': 'John Doe'},
                {'id': 2, 'name': 'Jane Smith'},
            ]
        """
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column['label'])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, column in enumerate(columns, start=1):
                value = row_data.get(column['key'], '')
                # Convert datetime to string
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_to_pdf(data: List[Dict[str, Any]], columns: List[Dict[str, str]], title: str = "Export") -> BytesIO:
        """
        Export data to PDF format.
        
        Args:
            data: List of dictionaries containing row data
            columns: List of dicts with 'key' and 'label' for each column
            title: Document title
            
        Returns:
            BytesIO: PDF file in memory
        """
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_paragraph = Paragraph(f"<b>{title}</b>", styles['Title'])
        elements.append(title_paragraph)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Prepare table data
        table_data = []
        
        # Headers
        headers = [column['label'] for column in columns]
        table_data.append(headers)
        
        # Rows
        for row_data in data:
            row = []
            for column in columns:
                value = row_data.get(column['key'], '')
                # Convert datetime to string
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M')
                row.append(str(value))
            table_data.append(row)
        
        # Create table
        table = Table(table_data)
        
        # Table styling
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Add footer with export date
        elements.append(Spacer(1, 0.5 * inch))
        footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, styles['Normal'])
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        return output
